import os
import re
import sys
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout,
    QLabel, QPushButton, QMessageBox, QFileDialog, QFrame
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from natsort import natsorted
import pandas as pd


class FileListGenerator(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.root_folder = ''
        self.output_file = ''
        self.organizations = ['과학기술사업화진흥원', '한국항공우주연구원', '국가안보실', '국가인권위원회', '국회도서관', '국회미래연구원', '국회사무처', '국회예산정책처', '국회입법조사처', '대통령경호처', '대통령비서실', '감사원', '고위공직자범죄수사처', '광주고등검찰청', '광주지방검찰청', '군사법원', '대검찰청', '대구고등검찰청', '대구지방검찰청', '대법원', '대전고등검찰청', '대전지방검찰청', '법무부', '법제처', '부산고등검찰청', '국가녹색기술연구소', '부산지방검찰청', '서울고등검찰청', '수원고등검찰청', '수원지방검찰청', '울산지방검찰청', '전주지방검찰청', '제주지방검찰청', '창원지방검찰청', '청주지방검찰청', '헌법재판소', '88관광개발주식회사', '개인정보보호위원회', '경제ㆍ인문사회연구회', '공정거래위원회', '국가보훈처', '국무조정실국무총리비서실', '국민권익위원회', '금융감독원', '금융위원회', '독립기념관', '서민금융진흥원', '신용보증기금', '예금보험공사', '중소기업은행', '한국공정거래조정원', '한국보훈복지의료공단', '한국산업은행', '한국소비자원', '한국자산관리공사', '한국주택금융공사', '관세청', '광주본부세관', '광주지방국세청', '국세청', '국제원산지정보원', '기획재정부', '대구본부세관', '대구지방국세청', '대전지방국세청', '부산본부세관', '부산지방국세청', '서울지방국세청', '인천지방국세청', '조달청', '중부지방국세청', '통계청', '한국수출입은행', '한국은행', '한국재정정보원', '한국조폐공사', '한국투자공사', '강릉원주대학교치과병원', '강원대학교', '강원대학교병원', '강원도교육청', '경기도교육청', '경북대학교', '경북대학교병원', '경북대학교치과병원', '경상국립대학교', '경상국립대학교병원', '경상남도교육청', '경상북도교육청', '광주광역시교육청', '교원소청심사위원회', '교육부', '국가교육위원회', '국가평생교육진흥원', '국립국제교육원', '국립특수교육원', '국사편찬위원회', '대구광역시교육청', '대전광역시교육청', '대한민국학술원사무국', '동북아역사재단', '부산광역시교육청', '부산대학교', '부산대학교병원', '부산대학교치과병원', '사립학교교직원연금공단', '서울과학기술대학교', '서울교육대학교', '서울대학교', '서울대학교병원', '서울대학교치과병원', '서울특별시교육청', '세종특별자치시교육청', '울산광역시교육청', '인천광역시교육청', '인천대학교', '전남대학교', '전남대학교병원', '전라남도교육청', '전라북도교육청', '전북대학교', '전북대학교병원', '제주대학교', '제주대학교병원', '제주특별자치도교육청', '중앙교육연수원', '충남대학교', '충남대학교병원', '충북대학교', '충북대학교병원', '충청남도교육청', '충청북도교육청', '한국고전번역원', '한국교원대학교', '한국교육시설안전원', '한국교육학술정보원', '한국교직원공제회', '한국대학교육협의회', '한국방송통신대학교', '한국사학진흥재단', '한국연구재단', '한국장학재단', '한국전문대학교육협의회', '한국학중앙연구원', '고등과학원', '과학기술연합대학원대학교', '과학기술인공제회', '과학기술일자리진흥원', '과학기술정보통신부', '광주과학기술원', '국가과학기술연구회', '국가과학기술인력개발원', '국가보안기술연구소', '국가수리과학연구소', '국립과천과학관', '국립광주과학관', '국립대구과학관', '국립부산과학관', '국립전파연구원', '국립중앙과학관', '기초과학연구원', '나노종합기술원', '녹색기술센터', '대구경북과학기술원', '방송문화진흥회', '방송통신심의위원회', '방송통신위원회', '별정우체국연금관리단', '세계김치연구소', '시청자미디어재단', '안전성평가연구소', '연구개발특구진흥재단', '우정사업본부', '우체국금융개발원', '우체국물류지원단', '우체국시설관리단', '울산과학기술원', '원자력안전위원회', '정보통신기획평가원', '정보통신산업진흥원', '중앙전파관리소', '한국건설기술연구원', '한국과학기술기획평가원', '한국과학기술단체총연합회', '한국과학기술연구원', '한국과학기술원', '한국과학기술정보연구원', '한국과학기술한림원', '한국과학영재학교', '한국과학창의재단', '한국교육방송공사', '한국기계연구원', '한국기초과학지원연구원', '한국나노기술원', '한국뇌연구원', '한국데이터산업진흥원', '한국방송공사', '한국방송광고진흥공사', '한국방송통신전파진흥원', '한국생명공학연구원', '한국생명기술연구원', '한국수력원자력', '한국식품연구원', '한국에너지기술연구원', '한국여성과학기술인육성재단', '한국연구재단', '한국우편사업진흥원', '한국원자력안전기술원', '한국원자력안전재단', '한국원자력연구원', '한국원자력의학원', '한국원자력통제기술원', '한국인터넷진흥원', '한국재료연구원', '한국전기연구원', '한국전자통신연구원', '한국지능정보사회진흥원', '한국지질자원연구원', '한국천문연구원', '한국철도기술연구원', '한국표준과학연구원', '한국한의학연구원', '한국공항우주연구원', '한국핵융합에너지연구원', '한국화학연구원', '남북교류협력지원협회', '민주평화통일자문회의', '북한이탈주민지원재단', '외교부', '재외동포재단', '통일부', '한ㆍ아프리카재단', '한국국제교류재단', '한국국제협력단', '5ㆍ18민주화운동진상규명조사위원회', '공군본부', '국방부', '방위사업청', '병무청', '육군본부', '지상군구성군사령부', '지방작전사령부', '합동참모본부', '해군본부', 'MG새마을금고중앙회', '경기남부경찰청', '경기도', '경기북부경찰청', '경상남도', '경상남도경찰청', '경상북도', '경상북도경찰청', '경찰공제회', '경찰청', '공무원연금공단', '광주광역시', '광주광역시경찰청', '대전광역시', '대전광역시경찰청', '대한소방공제회', '대한지방행정공제회', '도로교통공단', '민주화운동기념사업회', '바르게살기운동중앙협의회', '새마을운동중앙회', '서울경찰청', '서울특별시', '세종경찰청', '세종특별자치시', '소방산업공제조합', '소방청', '울산경찰청', '울산광역시', '인사혁신처', '일제강제동원피해자지원재단', '제주특별자치도', '제주특별자치도경찰청', '중앙선거관리위원회', '지방공기업평가원', '진실ㆍ화해를위한과거사정리위원회', '충청북도', '충청북도경찰청', '특수법인총포화약안전기술협회', '한국섬진흥원', '한국소방산업기술원', '한국소방시설협회', '한국소방안전원', '한국승강기안전공단', '한국자유총연맹', '한국지능정보사회진흥원', '한국지방세연구원', '한국지방재정공제회', '한국지방행정연구원', '한국지역정보개발원', '행정안전부',
                              '(재)국립극단', '(재)국립발레단', '한국생산기술연구원', '(재)국립심포니오케스트라', '(재)예술경영지원센터', '(재)한국공예ㆍ디자인문화진흥원', '(재)한국장애인문화예술원', '게임물관리위원회', '국립고궁박물관', '국립국악원', '국립국악중ㆍ고등학교', '국립국어원', '국립무형유산원', '국립문화재연구원', '국립민속박물관', '국립박물관문화재단', '국립아시아문화전당', '국립장애인도서관', '국립중앙극장', '국립중앙도서관', '국립중앙박물관', '국립한글박물관', '국립합창단', '국립해양문화재연구소', '국립현대미술관', '국외소재문화재재단', '국제방송교류재단', '궁능유적본부', '그랜드코리아레저㈜', '대한민국역사박물관', '대한장애인체육회', '대한체육회', '문화재정', '문화체육관광부', '사행산업통합감독위원회', '서울올림픽기념국민체육진흥공단', '세종장학재단', '스포츠윤리센터', '언론중재위원회', '영상물등급위원회', '영화진흥위원회', '예술원사무국', '예술의전당', '재단법인국악방송', '전통공연예술진흥재단', '태권도진흥재단', '한국관광공사', '한국도박문제예방치유원', '한국문학번역원', '한국문화관광연구원', '한국문화예술교육진흥원', '한국문화예술위원회', '한국문화예술회관연합회', '한국문화재재단', '한국문화정보원', '한국문화진흥주식회사', '한국언론진흥재단', '한국영상자료원', '한국예술인복지재단', '한국예술종합학교', '한국저작권보호원', '한국저작권위원회', '한국전통문화대학교', '한국정책방송원', '한국체육산업개발㈜', '한국출판문화산업진흥원', '한국콘텐츠진흥원', '해외문화홍보원', '현충사관리소', '가축위생방역지원본부', '극지연구소', '농림수산식품교육문화정보원', '농림식품기술기획평가원', '농림축산식품부', '농업정책보험금융원', '농업협동조합중앙회', '농촌진흥청', '부산항만공사', '산림조합중앙회', '산림청', '선박해양플랜트연구소', '수산업협동조합중앙회', '수협은행', '여수광양항만공사', '울산항만공사', '인천항만공사', '축산물품질평가원', '축산환경관리원', '한국농수산식품유통공사', '한국농어촌공사', '한국농업기술진흥원', '한국마사회', '한국산림복지진흥원', '한국수목원정원관리원', '한국수산자원공단', '한국식품산업클러스터진흥원', '한국어촌어항공단', '한국임업진흥원', '한국해양과학기술원', '한국해양교통안전공단', '한국해양수산연수원', '한국해양진흥공사', '해양경찰청', '해양수산과학기술진흥원', '해양수산부', '해양환경공단', '공영홈쇼핑', '기술보증기금', '대한무역투자진흥공사', '대한석탄공사', '산업통상자원부', '소상공인시장진흥공단', '신용보증재단중앙회', '재단법인장애인기업종합지원센터', '전략물자관리원', '전력거래소', '주식회사강원랜드', '중소기업기술정보진흥원', '중소기업유통센터', '중소벤처기업부', '중소벤처기업연구원', '중소벤처기업진흥공단', '창업진흥원', '특허청', '한국가스공사', '한국가스기술공사', '한국가스안전공사', '한국광해광업공단', '한국남동발전', '한국남부발전㈜', '한국동서발전', '한국디자인진흥원', '한국로봇산업진흥원', '한국무역보험공사', '한국발명진흥회', '한국벤처투자', '한국산업기술시험원', '한국산업기술진흥원', '한국산업기술평가관리원', '한국산업단지공단', '한국서부발전', '한국석유공사', '한국석유관리원', '한국세라믹기술원', '한국수력원자력㈜', '한국에너지공단', '한국에너지기술평가원', '한국에너지재단', '한국에너지정보문화재단', '한국원자력환경공단', '한국전기안전공사', '한국전력공사', '한국전력국제원자력대학원대학교', '한국전력기술주식회사', '한국제품안전관리원', '한국중부발전㈜', '한국지식재산보호원', '한국지식재산연구원', '한국지역난방공사', '한국탄소산업진흥원', '한국특허전략개발원', '한국특허정보원', '한전KDN㈜', '한전KPS', '한전엠씨에스㈜', '한전원자력연료㈜', '건강보험심사평가원', '국립암센터', '국립중앙의료원', '국민건강보험공단', '국민연금공단', '대구경북첨단의료산업진흥재단', '대학결핵협회', '대한적십자사', '보건복지부', '사회복지공동모금회', '식품안전정보원', '식품의약품안전처', '아동권리보장원', '오송첨단의료산업진흥재단', '의료기관평가인증원', '인구보건복지협회', '재단법인국가생명윤리정책원', '질병관리청', '한국건강증진개발원', '한국공공조직은행', '한국국제보건의료재단', '한국노인인력개발원', '한국마약퇴치운동본부', '한국보건복지인재원', '한국보건산업진흥원', '한국보건의료연구원', '한국보건의료인국가시험원', '한국보건의료정보원', '한국보육진흥원', '한국사회보장정보원', '한국사회복지협의회', '한국식품안전관리인증원', '한국의료기기안전정보원', '한국의료분쟁조정중재원', '한국의약품안전관리원', '한국자활복지개발원', '한국장기조직기증원', '한국장애인개발원', '한국한의약진흥원', '한국희귀필수의약품센터', '(재)차세대수치예보모델개발사업단', 'APEC기후센터', '건설근로자공제회', '경제사회노동위원회', '고용노동부', '고용노동부고객상담센터', '고용보험심사위원회', '광주지방고용노동청', '국가기상위성센터', '국가미세먼지정보센터', '국립공원공단', '국립기상과학원', '국립낙동강생물자원관', '국립생물자원관', '국립생태원', '국립야생동물질병관리원', '국립호남권생물자원관', '국립환경과학원', '국립환경인재개발원', '근로복지공단', '금강유역환경청', '금강홍수통제소', '기상기후인재개발원', '기상레이더센터', '기상청', '낙동강유역환경청', '노사발전재단', '대구지방고용노동청', '대구지방기상청', '대구지방환경청', '대전지방고용노동청', '부산지방고용노동청', '산업재해보상보험재심사위원회', '서울지방고용노동청', '수도권기상청', '수도권대기환경청', '수도권매립지관리공사', '수자원환경산업진흥㈜', '수치모델링센터', '영산강유역환경청', '영산강홍수통제소', '온실가스종합정보센터', '원주지방환경청', '전북지방환경청', '중부지방고용노동청', '중앙노동위원회', '중앙환경분쟁조정위원회', '최저임금위원회', '학교법인한국폴리텍', '한강유역환경청', '한강홍수통제소', '한국고용노동교육원', '한국고용정보원', '한국기상산업기술원', '한국기술교육대학교', '한국사회적기업진흥원', '한국산업안전보건공단', '한국산업인력공단', '한국상하수도협회', '한국수자원공사', '한국수자원조사기술원', '한국잡월드', '한국장애인고용공단', '한국환경공단', '한국환경산업기술원', '항공기상청', '화학물질안전원', '환경보전협회', '환경부', '건설기술교육원', '경기도', '공간정보품질관리원', '국가철도공단', '국립항공박물관', '국토교통과학기술진흥원', '국토교통부', '국토안전관리원', '대한건설기계안전관리원', '새만금개발공사', '새만금개발청', '서울특별시', '인천국제공항공사', '제주국제자유도시개발센터', '주식회사에스알', '주택관리공단㈜', '주택도시보증공사', '코레일관광개발㈜', '코레일네트웍스㈜', '코레일로지스㈜', '코레일유통㈜', '코레일테크㈜', '한국공항공사', '한국교통안전공단', '한국국토정보공사', '한국도로공사', '한국도로공사서비스', '한국부동산원', '한국철도공사', '한국해외인프라도시개발지원공사', '항공안전기술원', '행정중심복합도시건설청', '여성가족부', '한국건강가정진흥원', '한국양성평등교육진흥원', '한국여성인권진흥원', '한국청소년상담복지개발원', '한국청소년활동진흥원']
        self.names_21 = ['강기윤', '강대식', '강득구', '강민국', '강민정', '강병원', '강선우', '강성희', '강은미', '강준현', '강훈식', '고민정', '고영인', '고용진', '곽상도', '구자근', '권명호', '권성동', '권영세', '권은희', '권인숙', '권칠승', '기동민', '김경만', '김경협', '김교흥', '김근태', '김기현', '김남국', '김도읍', '김두관', '김미애', '김민기', '김민석', '김민철', '김병기', '김병욱', '김병욱', '김병주', '김상훈', '김상희', '김석기', '김선교', '김성원', '김성주', '김성환', '김수흥', '김승남', '김승수', '김승원', '김영배', '김영선', '김영식', '김영주', '김영진', '김영호', '김예지', '김용민', '김용판', '김웅', '김원이', '김윤덕', '김은혜', '김은희', '김의겸', '김정재', '김정호', '김종민', '김주영', '김진애', '김진표', '김철민', '김태년', '김태호', '김태흠', '김학용', '김한규', '김한정', '김형동', '김홍걸', '김회재', '김희곤', '김희국', '남인순', '노용호', '노웅래', '도종환', '류성걸', '류호정', '맹성규', '문정복', '문진석', '민병덕', '민형배', '민홍철', '박광온', '박대수', '박대출', '박덕흠', '박범계', '박병석', '박상혁', '박성민', '박성준', '박성중', '박수영', '박영순', '박완수', '박완주', '박용진', '박재호', '박정', '박정하', '박주민', '박진', '박찬대', '박형수', '박홍근', '배준영', '배진교', '배현진', '백종헌', '백혜련', '변재일', '서동용', '서범수', '서병수', '서삼석', '서영교', '서영석', '서일준', '서정숙', '설훈', '성일종', '소병철', '소병훈', '송갑석', '송기헌', '송석준', '송언석', '송영길', '송옥주', '송재호', '신동근', '신영대', '신원식', '신정훈', '신현영', '심상정', '안규백', '안민석', '안병길', '안철수', '안호영', '양경규', '양경숙', '양금희', '양기대', '양이원영', '양정숙',
                         '양향자', '어기구', '엄태영', '오기형', '오영환', '오영훈', '용혜인', '우상호', '우신구', '우원식', '위성곤', '유경준', '유기홍', '유동수', '유상범', '유의동', '유정주', '윤건영', '윤관석', '윤두현', '윤미향', '윤상현', '윤영덕', '윤영석', '윤영찬', '윤재갑', '윤재옥', '윤주경', '윤준병', '윤창현', '윤한홍', '윤호중', '윤후덕', '윤희숙', '이개호', '이광재', '이규민', '이낙연', '이달곤', '이동주', '이만희', '이명수', '이병훈', '이상민', '이상직', '이상헌', '이성만', '이소영', '이수진', '이수진', '이양수', '이영', '이용', '이용빈', '이용선', '이용우', '이용호', '이원욱', '이원택', '이은주', '이인선', '이인영', '이자스민', '이장섭', '이재명', '이재정', '이정문', '이종배', '이종성', '이주환', '이채익', '이철규', '이탄희', '이태규', '이학영', '이해식', '이헌승', '이형석', '인재근', '임병헌', '임오경', '임이자', '임종성', '임호선', '장경태', '장동혁', '장제원', '장철민', '장혜영', '전봉민', '전용기', '전재수', '전주혜', '전해철', '전혜숙', '정경희', '정동만', '정성호', '정우택', '정운천', '정일영', '정점식', '정정순', '정진석', '정찬민', '정청래', '정춘숙', '정태호', '정필모', '정희용', '조경태', '조명희', '조수진', '조승래', '조오섭', '조은희', '조응천', '조정식', '조정훈', '조태용', '조해진', '주철현', '주호영', '지성호', '진선미', '진성준', '천준호', '최강욱', '최기상', '최승재', '최연숙', '최영희', '최인호', '최재형', '최종윤', '최춘식', '최형두', '최혜영', '추경호', '태영호', '하영제', '하태경', '한기호', '한무경', '한병도', '한정애', '한준호', '허숙정', '허영', '허은아', '허종식', '홍기원', '홍문표', '홍석준', '홍성국', '홍영표', '홍익표', '홍정민', '홍준표', '황보승희', '황운하', '황희']
        self.file_attach = ['붙임', '별도', '별첨']
        self.file_answer = ['답변서', '답변자료', '질의 답변', '질의답변']
        self.file_require = ['공통요구', '요구자료', '자료요구', '위원 요구', '감사 요구', '감사요구']

    def select_root_folder(self):
        self.root_folder = QFileDialog.getExistingDirectory(self, '폴더를 선택하세요')
        if self.root_folder:
            self.label_selected_root_folder.setText(
                f'선택한 폴더: {self.root_folder}')
        else:
            self.label_selected_root_folder.setText('선택한 폴더: 선택하지 않음')

        if self.root_folder and self.output_file:
            self.generate_button.setEnabled(True)

    def select_output_dir(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.FileMode.AnyFile)
        file_dialog.setNameFilter("Excel 파일 (*.xlsx)")
        file_dialog.setViewMode(QFileDialog.ViewMode.Detail)

        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles()
            self.output_file = selected_files[0]
            self.label_selected_output_dir.setText(
                f'선택한 파일: {self.output_file}')
        else:
            self.label_selected_output_dir.setText('선택한 파일: 선택하지 않음')

        if self.root_folder and self.output_file:
            self.generate_button.setEnabled(True)

    def initUI(self):
        self.setWindowTitle('국정감사 메타데이터 생성기')
        self.resize(300, 200)

        layout = QVBoxLayout()

        self.folder_path_label = QLabel('폴더를 선택하세요:')
        self.folder_path_input = QPushButton('폴더 선택')
        self.folder_path_input.clicked.connect(self.select_root_folder)
        self.label_selected_root_folder = QLabel('선택한 폴더: 선택하지 않음')
        layout.addWidget(self.folder_path_label)
        layout.addWidget(self.folder_path_input)
        layout.addWidget(self.label_selected_root_folder)

        self.output_path_label = QLabel('저장할 엑셀 파일 경로:')
        self.output_path_input = QPushButton('파일 선택')
        self.output_path_input.clicked.connect(self.select_output_dir)
        self.label_selected_output_dir = QLabel('선택한 경로: 선택하지 않음')
        layout.addWidget(self.output_path_label)
        layout.addWidget(self.output_path_input)
        layout.addWidget(self.label_selected_output_dir)

        self.separator = QFrame()
        self.separator.setFrameShape(QFrame.Shape.HLine)
        self.separator.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(self.separator)

        self.generate_button = QPushButton('엑셀 생성')
        self.generate_button.clicked.connect(self.generate_file_list)
        self.generate_button.setDefault(True)
        self.generate_button.setEnabled(False)
        layout.addWidget(self.generate_button)

        self.setLayout(layout)

    def generate_file_list(self):
        parent_folder = self.root_folder
        output_file = self.output_file

        if not os.path.isdir(parent_folder):
            QMessageBox.warning(self, '경로 오류', '유효하지 않은 폴더 경로입니다.')
            return

        # 최상위 폴더명 가져오기
        top_level_folder = os.path.basename(parent_folder)
        grandparent_folder = os.path.dirname(parent_folder)

        # 파일 리스트 초기화
        file_list = []

        # parent_folder를 기준으로 모든 파일을 탐색
        for root, _, files in os.walk(parent_folder):
            # 파일을 자연 정렬하여 순회
            for file in natsorted(files):
                file_path = os.path.join(root, file)  # 파일 경로 생성

                # 파일의 부모 폴더명 가져오기
                relative_path = os.path.relpath(file_path, grandparent_folder)
                first_folder_name = relative_path.split(
                    os.sep)[1]  # 첫 번째 서브폴더명 가져오기

                # 파일 정보를 딕셔너리로 추가
                file_list.append({
                    '위원회': top_level_folder,
                    '피감기관': first_folder_name,
                    'FILE_NAME': file,  # 파일명
                    '실제 경로': relative_path,  # 실제 경로
                })

        # DataFrame 생성
        df = pd.DataFrame(file_list)

        # 출력 파일이 존재하지 않는 경우 새로운 워크북 생성
        add_extension_filename = output_file + '.xlsx'

        if not os.path.exists(output_file) and not os.path.exists(add_extension_filename):
            if not output_file.endswith('.xlsx'):
                output_file = add_extension_filename
            wb = Workbook()
            ws = wb.active

            # 헤더 추가
            headers = ['위원회', '피감기관', 'BOOK_ID', 'SEQNO', 'FILE_NAME',
                       '국정감사 파일명', '위원', '질의', 'REALFILE_NAME', '실제 경로', '문서 종류']
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # 첫 번째 행의 셀 색상 설정
            fill_color = PatternFill(start_color='4f81bd',
                                     end_color='4f81bd', fill_type='solid')
            for col in range(1, 12):
                ws.cell(row=1, column=col).fill = fill_color
        else:
            # 기존 파일 불러오기
            try:
                if os.path.exists(output_file):
                    wb = load_workbook(output_file)
                elif os.path.exists(add_extension_filename):
                    wb = load_workbook(add_extension_filename)
            except Exception as e:
                QMessageBox.warning(self, '엑셀 파일 읽기 오류', f'{e} 엑셀 파일 확장자 오류')
                return
            ws = wb.active

        # 마지막 행 인덱스 확인
        last_row = ws.max_row

        # DataFrame의 각 행을 엑셀에 추가
        for index, row in df.iterrows():
            blank = str(row['위원회']).find(' ')
            if blank != -1:
                tmp_org = str(row['위원회'])[blank+1:]
            else:
                tmp_org = row['위원회']
            ws.cell(row=last_row + index + 1,
                    column=1, value=tmp_org)  # 위원회
            # 정규표현식으로 피감기관 검색
            pattern = '|'.join(
                rf'{re.escape(org)}' for org in self.organizations)
            matches = re.search(pattern, row['피감기관'])
            if matches:
                ws.cell(row=last_row + index + 1, column=2,
                        value=matches[0])  # 피감기관
            else:
                ws.cell(row=last_row + index + 1, column=2,
                        value=row['피감기관'])  # 피감기관
            # 정규표현식으로 위원 검색
            pattern2 = '|'.join(
                rf'{re.escape(org)}' for org in self.names_21)
            matches2 = re.search(pattern2, row['실제 경로'])
            if matches2:
                if matches2[0] == '이용' and tmp_org != '문화체육관광위원회':
                    if len(matches2.groups()) > 1 and matches2[1]:
                        ws.cell(row=last_row + index + 1, column=7,
                                value=matches2[-1] + ' 위원')  # 위원
                    else:
                        ws.cell(row=last_row + index + 1,
                                column=7, value=None)  # 위원
                else:
                    ws.cell(row=last_row + index + 1, column=7,
                            value=matches2[0] + ' 위원')  # 위원
            else:
                # 실제 경로 폴더명에서 위원명이 검출되지 않을경우 공백
                ws.cell(row=last_row + index + 1, column=7, value=None)
            ws.cell(row=last_row + index + 1, column=9,
                    value=row['FILE_NAME'])  # 파일명
            ws.cell(row=last_row + index + 1, column=10,
                    value=row['실제 경로'])  # 실제 경로

            result_file = self.search_in_row(row['FILE_NAME'])
            result_dir = self.search_in_row(row['실제 경로'])
            if result_file == 1:
                ws.cell(row=last_row + index + 1, column=11, value='별도제출자료')
            elif result_file == 2:
                ws.cell(row=last_row + index + 1,
                        column=11, value='서면 질의 답변자료')
            elif result_file == 3:
                ws.cell(row=last_row + index + 1, column=11, value='국정감사 요구자료')
            else:
                if result_dir == 1:
                    ws.cell(row=last_row + index + 1,
                            column=11, value='별도제출자료')
                elif result_dir == 2:
                    ws.cell(row=last_row + index + 1,
                            column=11, value='서면 질의 답변자료')
                elif result_dir == 3:
                    ws.cell(row=last_row + index + 1,
                            column=11, value='국정감사 요구자료')
                else:
                    ws.cell(row=last_row + index + 1, column=11, value='기타')

        # 변경 사항 저장
        if os.path.exists(output_file) or output_file.endswith('.xlsx'):
            wb.save(output_file)
        else:
            wb.save(add_extension_filename)

        # 완료 메시지 출력
        QMessageBox.information(self, '완료', f'파일 정보가 {output_file}에 저장되었습니다.')

    def search_in_row(self, row):
        pattern_attach = '|'.join(
            rf'{re.escape(org)}' for org in self.file_attach)
        matches_attach = re.search(pattern_attach, row)
        pattern_answer = '|'.join(
            rf'{re.escape(org)}' for org in self.file_answer)
        matches_answer = re.search(pattern_answer, row)
        pattern_require = '|'.join(
            rf'{re.escape(org)}' for org in self.file_require)
        matches_require = re.search(pattern_require, row)

        if matches_attach:
            return 1
        elif matches_answer:
            return 2
        elif matches_require:
            return 3
        else:
            return 4


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = FileListGenerator()
    window.show()
    sys.exit(app.exec())
